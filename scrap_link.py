import requests
from bs4 import BeautifulSoup
import pandas as pd
import os
import re
from openpyxl import load_workbook


sitemap_file = 'resource/links.xlsx'
df_links = pd.read_excel(sitemap_file)
urls = df_links['URL'].dropna().tolist()

for url in urls:
    r = requests.get(url)
    soup = BeautifulSoup(r.content, 'html.parser')

    # استخراج بخش محتوای اصلی (قبل از سوالات متداول)
    content_paragraphs = []
    min_word_count = 15
    found_faq_header = False

    for element in soup.find_all(['p', 'h3']):
        if element.name == 'h3' and 'سوالات متداول' in element.text:
            found_faq_header = True
            break
        elif element.name == 'p' and not found_faq_header:
            text = element.text.strip()
            words = text.split()
            if len(words) >= min_word_count:
                if len(words) > 500:
                    print(f"[مشکوک] پاراگراف طولانی: {len(words)} کلمه\n{text[:200]}...\n")
                    # continue  
                content_paragraphs.append(text)

    print(f"تعداد پاراگراف‌های محتوای اصلی واجد شرایط: {len(content_paragraphs)}")

    # ذخیره محتوای اصلی در فایل Excel
    output_directory = 'result'
    os.makedirs(output_directory, exist_ok=True)
    content_output_file = os.path.join(output_directory, 'data.xlsx')
    df_content = pd.DataFrame({'متن پاراگراف': content_paragraphs})

    if os.path.exists(content_output_file):
        book = load_workbook(content_output_file)
        start_row = book.active.max_row
        with pd.ExcelWriter(content_output_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            df_content.to_excel(writer, sheet_name='Sheet1', startrow=start_row, header=False, index=False)
        print(f"محتوای جدید به فایل '{content_output_file}' اضافه شد.")
    else:
        df_content.to_excel(content_output_file, index=False)
        print(f"محتوای پاراگراف‌های اصلی در فایل '{content_output_file}' ذخیره شد.")


    # استخراج بخش سوالات متداول
    faq_items = []
    faq_header = soup.find('h3', string='سوالات متداول')

    if faq_header:
        current_element = faq_header.find_next_sibling()
        current_question = None
        while current_element:
            if current_element.name == 'h4':
                question_text = current_element.text.strip()
                current_question = re.sub(r'^\d+- ✔️\s*', '', question_text)
            elif current_element.name == 'p' and current_question:
                answer_text = current_element.text.strip()
                answer_without_checkmark = answer_text.replace('✔️', '').strip()
                answer = answer_without_checkmark.replace('با توجه به متن مقاله، ', '').strip()
                faq_items.append({'question': current_question, 'answer': answer})
                current_question = None
            current_element = current_element.find_next_sibling()

        # ذخیره سوالات متداول در فایل Excel
        faq_output_file = os.path.join(output_directory, 'qa.xlsx')
        df_faq = pd.DataFrame(faq_items)

        if os.path.exists(faq_output_file):
            book = load_workbook(faq_output_file)
            start_row = book.active.max_row
            with pd.ExcelWriter(faq_output_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                df_faq.to_excel(writer, sheet_name='Sheet1', startrow=start_row, header=False, index=False)
            print(f"سوالات متداول جدید به فایل '{faq_output_file}' اضافه شد.")
        else:
            df_faq.to_excel(faq_output_file, index=False)
            print(f"سوالات متداول در فایل '{faq_output_file}' ذخیره شد.")


    else:
        print("بخش 'سوالات متداول' در صفحه یافت نشد.")